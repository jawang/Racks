import openpyxl as xl

# type constants
OLD = 0
NEW = 1

class Routers:
    def __init__(self):
        # import excel data
        racks = xl.load_workbook('racks.xlsx')
        rackdata = racks.get_sheet_by_name('racks')

        # set up old groups
        self.groups = [Group(OLD) for i in range(5)]
        for i in range(5):
            for j in range(0,4):
                self.groups[i].boxes[j].rackname = \
                    str(rackdata.cell(row=i+1,column=1).value)
            for j in range(4,8):
                self.groups[i].boxes[j].rackname = \
                    str(rackdata.cell(row=i+1,column=2).value)
            for j in range(8,10):
                self.groups[i].boxes[j].rackname = \
                    str(rackdata.cell(row=i+1,column=3).value)
            for j in range(10,13):
                self.groups[i].boxes[j].rackname = \
                    str(rackdata.cell(row=i+1,column=4).value)

        # set up new groups
        self.groups += [Group(NEW) for i in range(2)]
        for i in range(5,7):
            for j in range(0,5):
                self.groups[i].boxes[j].rackname = \
                    str(rackdata.cell(row=i+1,column=2).value)
            for j in range(5,7):
                self.groups[i].boxes[j].rackname = \
                    str(rackdata.cell(row=i+1,column=3).value)
            self.groups[i].boxes[j+1].rackname = \
                str(rackdata.cell(row=i+1,column=1).value)

    def find(self,inputnum,outputnum):
        # find the group by output
        if outputnum in range(1,641):
            gindex = (outputnum-1)/128
            group = self.groups[gindex]
            
            # find the box by input
            if inputnum in range(1,641):
                bindex = (inputnum-1)/64 
                box = group.boxes[bindex]
            
            elif inputnum in range(641,1025):
                bindex = (inputnum-640-1)/128+10
                box = group.boxes[bindex]
        elif outputnum in range(641,1025):
            gindex = (outputnum-513-1)/256+5
            group = self.groups[gindex]
            bindex = (inputnum-1)/128
            box = group.boxes[bindex]
        else:
            print 'Invalid output '+str(outputnum)
            return

        return [box.rackname, gindex+1, bindex+1]
    
class Group:
    def __init__(self,gtype):
        if gtype==OLD:
            self.type = 'OLD'
            self.boxes = [Box(OLD,OLD) for i in range(10)]
            self.boxes += [Box(OLD,NEW) for i in range(3)]
        elif gtype==NEW:
            self.type = 'NEW'
            self.boxes = [Box(NEW,NEW) for i in range(8)]
        else:
            print 'Invalid gtype '+str(gtype)

'''
class Rack:
    def __init__(self, name):

'''       

class Box:
    def __init__(self,gtype,btype):
        self.rackname = ''
        if gtype==OLD:
            if btype==OLD:
                self.type = 'TRS BOX'
                self.chassis = [Chassis(OLD,OLD) \
                                for i in range(3)]
            elif btype==NEW:
                self.type = 'ECLIPSE BOX'
                self.chassis = [Chassis(OLD,NEW)]
            else:
                print 'Invalid btype '+str(btype)
        elif gtype==NEW:
            if btype==NEW:
                self.type = 'ECLIPSE BOX'
                self.chassis = [Chassis(NEW,NEW)]
            else:
                print 'Invalid pair ('+str(gtype)+','+str(btype)+')'
        else:
            print 'Invalid gtype '+str(gtype)

class Chassis:
    def __init__(self,gtype,btype):
        if gtype==OLD:
            if btype==OLD:
                self.type = 'TRS'
            elif btype==NEW:
                self.type = 'ECLIPSE'
            else:
                print 'Invalid btype '+str(btype)
        elif gtype==NEW:
            if btype==NEW:
                self.type = 'ECLIPSE'
            else:
                print 'Invalid pair ('+str(gtype)+','+str(btype)+')'
        else:
            print 'Invalid gtype '+str(gtype)
